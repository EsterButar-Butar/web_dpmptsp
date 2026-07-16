import openpyxl
import os

def check_file(filename):
    print(f"=== File: {filename} ===")
    if not os.path.exists(filename):
        print("File does not exist.")
        return
    try:
        wb = openpyxl.load_workbook(filename, read_only=True)
        print("Sheet Names:", wb.sheetnames)
        for name in wb.sheetnames:
            sheet = wb[name]
            # Read first few rows
            rows = list(sheet.iter_rows(max_row=3, values_only=True))
            print(f"  Sheet '{name}':")
            if rows:
                print("    Header:", rows[0])
                if len(rows) > 1:
                    print("    Row 2:", rows[1])
                if len(rows) > 2:
                    print("    Row 3:", rows[2])
    except Exception as e:
        print("Error:", e)
    print()

check_file('test_import.xlsx')
check_file('test_import_sektoral.xlsx')
